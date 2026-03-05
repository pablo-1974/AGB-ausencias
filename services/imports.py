# services/imports.py
from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Tuple

import openpyxl
from fastapi import HTTPException, UploadFile
from sqlalchemy import select, delete
from sqlalchemy.orm import Session

from models import Teacher, GuardDuty, TeachingSlot, DutyType


# ============================================================
# IMPORTAR LISTADO DE PROFESORES  —  Opción B (omitir duplicados)
#   - Espera una hoja con columnas: nombre | email (en ese orden)
#   - Inserta los nuevos
#   - Omite (skip) los que ya existan por email (NO actualiza)
#   - Devuelve {'inserted': X, 'skipped': Y}
# ============================================================

async def import_teachers_file(upload_file: UploadFile, db: Session) -> Dict[str, int]:
    """
    Importa profesores desde un Excel (.xlsx/.xls) con columnas:
    nombre | email

    Regla:
      - si el email ya existe: OMITE la fila (opción B)
      - si es nuevo: crea Teacher(name, email, is_current=True)

    Retorno:
      {'inserted': X, 'skipped': Y}
    """
    if not upload_file.filename:
        raise HTTPException(status_code=400, detail="No se ha enviado ningún archivo.")
    if not (upload_file.filename.endswith(".xlsx") or upload_file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls.")

    # Cargar el libro desde el stream del UploadFile
    try:
        # Nota: openpyxl admite file-like objects; usamos .file directamente
        wb = openpyxl.load_workbook(upload_file.file, data_only=True)
    except Exception as e:
        # Si por alguna razón el stream no es re‑leíble, plan B (memoria):
        upload_file.file.seek(0)
        try:
            data = upload_file.file.read()
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        except Exception as e2:
            raise HTTPException(status_code=400, detail=f"No se pudo abrir el Excel: {e2}") from e

    ws = wb.active
    # Validación mínima de estructura
    if ws.max_row < 2 or ws.max_column < 2:
        raise HTTPException(status_code=400, detail="La hoja parece vacía o sin columnas 'nombre|email'.")

    inserted = 0
    skipped = 0

    # Recorremos desde la fila 2 (cabecera en la 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        # Se espera: nombre | email (solo las dos primeras columnas)
        name, email = (row + (None, None))[:2]  # por si faltan celdas
        if not name or not email:
            # Fila incompleta → omitimos silenciosamente
            continue

        email_norm = str(email).strip().lower()
        name_norm = str(name).strip()

        if not email_norm or not name_norm:
            continue

        # ¿Ya existe por email? → OMITE (skip) y continúa
        exists = db.scalar(select(Teacher.id).where(Teacher.email == email_norm))
        if exists:
            skipped += 1
            continue

        # Nuevo profesor
        prof = Teacher(
            name=name_norm,
            email=email_norm,
            is_current=True,  # tu campo actual
        )
        db.add(prof)
        inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped}


# ============================================================
# IMPORTAR HORARIO (CLASES + GUARDIAS) — SIN REUNIONES
#   - Borra GuardDuty y TeachingSlot
#   - Espera columnas: nombre | día | franja | grupo | aula | materia
#   - "G AULA" / "G RECREO" → guardias
#   - Resto → clase normal
# ============================================================

async def import_schedule_file(upload_file: UploadFile, db: Session) -> None:
    """
    Importa el horario (sin reuniones) desde un Excel.
    Estructura esperada de cada fila (desde la 2):
      nombre | día | franja | grupo | aula | materia

    Reglas:
      - "G AULA" o "G RECREO" en la columna 'grupo' → GuardDuty
      - En otro caso → TeachingSlot
    """
    if not upload_file.filename:
        raise HTTPException(status_code=400, detail="No se ha enviado ningún archivo.")
    if not (upload_file.filename.endswith(".xlsx") or upload_file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls.")

    # Carga del libro
    try:
        wb = openpyxl.load_workbook(upload_file.file, data_only=True)
    except Exception:
        upload_file.file.seek(0)
        data = upload_file.file.read()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)

    ws = wb.active

    # Borrar datos antiguos
    db.execute(delete(GuardDuty))
    db.execute(delete(TeachingSlot))
    db.commit()

    # Mapeo de días y normalización de franja
    daymap = {
        "lunes": 0, "martes": 1,
        "miércoles": 2, "miercoles": 2,
        "jueves": 3, "viernes": 4
    }

    def normalize_slot(s) -> str:
        s = str(s).strip().lower()
        mapping = {
            "1ª": "1", "1": "1",
            "2ª": "2", "2": "2",
            "3ª": "3", "3": "3",
            "recreo": "RECREO",
            "4ª": "4", "4": "4",
            "5ª": "5", "5": "5",
            "6ª": "6", "6": "6",
        }
        return mapping.get(s, s.upper())

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        # nombre | día | franja | grupo | aula | materia
        name, day, slot, group, room, subject = (row + (None,)*6)[:6]

        if not name or not day or not slot:
            continue

        # Buscar profesor por nombre exacto
        teacher = db.execute(
            select(Teacher).where(Teacher.name == str(name).strip())
        ).scalar_one_or_none()

        if not teacher:
            # Si no existe el profesor, omitimos la fila
            continue

        weekday = daymap.get(str(day).lower().strip())
        if weekday is None:
            continue

        slot_norm = normalize_slot(slot)

        # Guardias (G AULA / G RECREO)
        group_str = str(group).strip().upper() if group else ""
        if group_str in ("G AULA", "G RECREO"):
            duty_type = DutyType.AULA if group_str == "G AULA" else DutyType.RECREO
            db.add(GuardDuty(
                teacher_id=teacher.id,
                weekday=weekday,
                slot=slot_norm,
                type=duty_type
            ))
            continue

        # Clases normales
        db.add(TeachingSlot(
            teacher_id=teacher.id,
            weekday=weekday,
            slot=slot_norm,
            group=str(group).strip() if group else "",
            room=str(room).strip() if room else "",
            subject=str(subject).strip() if subject else ""
        ))

    db.commit()
